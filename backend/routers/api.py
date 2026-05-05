"""API endpoints for HTMX interactions, data operations, charts, and exports."""
from fastapi import APIRouter, Request, Query, Form, UploadFile, File, Depends
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from ..services.cache import sync_all, get_sync_info, calc_shipping
from ..services.sheets import append_order_to_sheet, parse_money
from ..services.tracking import parse_tracking_info
from ..services.buffer import save_to_buffer, get_buffer_count, flush_buffer
from ..models.database import get_db
from ..config import DON_RATE_PER_KG, DON_RATE_PER_M3, DON2_RATE_PER_KG, CREDENTIALS_DIR, BASE_DIR
from ..auth import verify_user, verify_admin
import json
import os
import shutil
import io
from datetime import datetime, timedelta
from unidecode import unidecode

router = APIRouter(tags=["api"])


def js_escape(s: str) -> str:
    """Escape string for safe embedding in JS onclick handler."""
    if not s:
        return ""
    return (s
        .replace("\\", "\\\\")
        .replace("'", "\\'")
        .replace('"', '\\"')
        .replace("\n", "\\n")
        .replace("\r", "\\r")
        .replace("<", "&lt;")
        .replace(">", "&gt;"))

@router.post("/sync")
async def sync_data():
    """Sync from Google Sheets to local cache."""
    try:
        count = await sync_all()
        info = await get_sync_info()
        return HTMLResponse(f"""
        <div class="flex items-center gap-2 text-green-600">
            <svg class="w-5 h-5" fill="currentColor" viewBox="0 0 20 20">
                <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-9.293a1 1 0 00-1.414-1.414L9 10.586 7.707 9.293a1 1 0 00-1.414 1.414l2 2a1 1 0 001.414 0l4-4z" clip-rule="evenodd"/>
            </svg>
            <span>Đã đồng bộ {count} đơn hàng • {info['customer_count']} khách hàng</span>
        </div>
        """)
    except Exception as e:
        return HTMLResponse(f"""
        <div class="flex items-center gap-2 text-red-600">
            <svg class="w-5 h-5" fill="currentColor" viewBox="0 0 20 20">
                <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zM8.707 7.293a1 1 0 00-1.414 1.414L8.586 10l-1.293 1.293a1 1 0 101.414 1.414L10 11.414l1.293 1.293a1 1 0 001.414-1.414L11.414 10l1.293-1.293a1 1 0 00-1.414-1.414L10 8.586 8.707 7.293z" clip-rule="evenodd"/>
            </svg>
            <span>Lỗi: {str(e)}</span>
        </div>
        """, status_code=500)

@router.get("/search-customer")
async def search_customer(q: str = Query("", alias="q")):
    """Quick customer search for form auto-fill."""
    if not q or len(q) < 2:
        return HTMLResponse("")

    db = await get_db()
    try:
        q_clean = q.replace(" ", "")
        q_ascii = unidecode(q)
        results = await db.execute_fetchall(
            """SELECT customer_name, customer_phone, customer_address,
                COUNT(*) as order_count
            FROM orders
            WHERE customer_phone LIKE ? OR customer_name LIKE ?
            GROUP BY customer_phone
            ORDER BY order_count DESC
            LIMIT 10""",
            (f"%{q_clean}%", f"%{q}%")
        )

        # If no results, try diacritics-free search
        if not results and q_ascii != q:
            results = await db.execute_fetchall(
                """SELECT customer_name, customer_phone, customer_address,
                    COUNT(*) as order_count
                FROM orders
                WHERE customer_phone LIKE ? OR customer_name LIKE ?
                GROUP BY customer_phone
                ORDER BY order_count DESC
                LIMIT 10""",
                (f"%{q_clean}%", f"%{q_ascii}%")
            )

        if not results:
            return HTMLResponse("")

        items = ""
        for r in results:
            items += f"""
            <div class="px-3 py-2 hover:bg-blue-50 cursor-pointer flex justify-between items-center"
                 hx-on:click="fillCustomer('{js_escape(r[0])}', '{js_escape(r[1])}', '{js_escape(r[2])}')"
                 hx-on:mouseenter="this.classList.add('bg-blue-50')"
                 hx-on:mouseleave="this.classList.remove('bg-blue-50')">
                <div>
                    <div class="font-medium text-gray-900">{r[0]}</div>
                    <div class="text-sm text-gray-500">{r[1]} • {r[3]} đơn</div>
                </div>
                <div class="text-xs text-gray-400">{r[2][:30] if r[2] else ''}</div>
            </div>"""

        return HTMLResponse(f"""
        <div class="absolute z-50 w-full bg-white border border-gray-200 rounded-lg shadow-lg mt-1 max-h-60 overflow-y-auto">
            {items}
        </div>
        """)
    finally:
        await db.close()


@router.get("/customer-history")
async def customer_history(phone: str = Query("", alias="phone")):
    """Get recent orders for a customer phone - Smart Form feature."""
    if not phone or len(phone) < 3:
        return HTMLResponse("")

    db = await get_db()
    try:
        q_clean = phone.replace(" ", "")
        orders = await db.execute_fetchall(
            """SELECT o.id, o.order_date, o.customer_name, o.sheet_type,
                o.total_price, o.deposit, o.remaining, o.status,
                o.tracking_cn, o.tracking_vn,
                GROUP_CONCAT(oi.product_name, ' | ') as products
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            WHERE o.customer_phone LIKE ?
            GROUP BY o.id
            ORDER BY o.row_start DESC
            LIMIT 3""",
            (f"%{q_clean}%",)
        )

        if not orders:
            return HTMLResponse("")

        rows = ""
        for o in orders:
            status_badge = ""
            if o[7]:
                color = "green" if "hoàn" in (o[7] or "").lower() or "xong" in (o[7] or "").lower() else "blue"
                status_badge = f'<span class="px-1.5 py-0.5 rounded text-[10px] bg-{color}-100 text-{color}-700">{o[7]}</span>'

            rows += f"""
            <tr class="border-b border-gray-100 last:border-0">
                <td class="px-2 py-1.5 text-xs text-gray-500">{o[1] or ''}</td>
                <td class="px-2 py-1.5">
                    <a href="/don-hang/{o[0]}" class="text-xs text-primary-600 hover:underline">{o[10][:20] if o[10] else o[8][:15] if o[8] else '-'}</a>
                </td>
                <td class="px-2 py-1.5 text-xs max-w-[120px] truncate">{o[11] or ''}</td>
                <td class="px-2 py-1.5 text-xs text-right font-medium">{o[4]:,.0f} đ</td>
                <td class="px-2 py-1.5">{status_badge}</td>
            </tr>"""

        return HTMLResponse(f"""
        <div class="mt-2 bg-amber-50 border border-amber-200 rounded-lg p-3">
            <div class="flex items-center gap-2 mb-2">
                <span class="text-amber-700 text-xs font-medium">📋 3 đơn gần nhất của {orders[0][2]}</span>
                <span class="text-[10px] text-amber-500">({len(orders)} đơn)</span>
            </div>
            <table class="w-full">
                <thead>
                    <tr class="text-[10px] text-amber-600 uppercase">
                        <th class="px-2 py-1 text-left">Ngày</th>
                        <th class="px-2 py-1 text-left">Mã</th>
                        <th class="px-2 py-1 text-left">SP</th>
                        <th class="px-2 py-1 text-right">Giá</th>
                        <th class="px-2 py-1 text-left">TT</th>
                    </tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        </div>
        """)
    finally:
        await db.close()


@router.get("/search-tracking")
async def search_tracking(q: str = Query("", alias="q")):
    """Search by tracking number."""
    if not q or len(q) < 3:
        return HTMLResponse("")

    db = await get_db()
    try:
        results = await db.execute_fetchall(
            """SELECT o.id, o.customer_name, o.customer_phone, o.tracking_cn, o.tracking_vn,
                o.status, o.sheet_type, GROUP_CONCAT(oi.product_name, ' | ') as products
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            WHERE o.tracking_cn LIKE ? OR o.tracking_vn LIKE ?
            GROUP BY o.id LIMIT 20""",
            (f"%{q}%", f"%{q}%")
        )

        if not results:
            return HTMLResponse('<div class="p-3 text-gray-500 text-sm">Không tìm thấy</div>')

        items = ""
        for r in results:
            status_color = "green" if r[5] else "gray"
            items += f"""
            <a href="/don-hang/{r[0]}" class="block px-3 py-2 hover:bg-blue-50 border-b border-gray-100">
                <div class="flex justify-between">
                    <span class="font-medium">{r[1] or 'N/A'}</span>
                    <span class="text-xs px-2 py-0.5 rounded bg-{status_color}-100 text-{status_color}-700">{r[6]}</span>
                </div>
                <div class="text-sm text-gray-500">TQ: {r[3] or '-'} → VN: {r[4] or '-'}</div>
                <div class="text-xs text-gray-400">{r[7] or ''}</div>
            </a>"""

        return HTMLResponse(f"""
        <div class="absolute z-50 w-full bg-white border border-gray-200 rounded-lg shadow-lg mt-1 max-h-80 overflow-y-auto">
            {items}
        </div>
        """)
    finally:
        await db.close()

@router.get("/order-items/{order_id}")
async def get_order_items(order_id: int):
    """Get order items for detail view."""
    db = await get_db()
    try:
        items = await db.execute_fetchall(
            "SELECT * FROM order_items WHERE order_id = ?", (order_id,)
        )
        if not items:
            return HTMLResponse('<div class="text-gray-400 text-sm">Không có sản phẩm</div>')

        rows = ""
        for item in items:
            rows += f"""
            <tr class="border-b">
                <td class="px-3 py-2">{item[3] or ''}</td>
                <td class="px-3 py-2 text-right">{item[4]:.1f} kg</td>
                <td class="px-3 py-2 text-right">{item[5]:.2f} m³</td>
                <td class="px-3 py-2 font-mono text-sm">{item[6] or ''}</td>
                <td class="px-3 py-2 text-right">{item[8]:,.0f} đ</td>
            </tr>"""

        return HTMLResponse(f"""
        <table class="w-full text-sm">
            <thead class="bg-gray-50">
                <tr>
                    <th class="px-3 py-2 text-left">Sản phẩm</th>
                    <th class="px-3 py-2 text-right">Khối lượng</th>
                    <th class="px-3 py-2 text-right">Thể tích</th>
                    <th class="px-3 py-2 text-left">Vận đơn TQ</th>
                    <th class="px-3 py-2 text-right">Giá</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        """)
    finally:
        await db.close()

@router.post("/create-order")
async def create_order(
    sheet_type: str = Form(...),
    order_date: str = Form(""),
    customer_name: str = Form(...),
    customer_phone: str = Form(""),
    customer_address: str = Form(""),
    product_name: str = Form(""),
    note: str = Form(""),
    total_price: int = Form(0),
    deposit: int = Form(0),
):
    """Create new order and write to Google Sheets."""

    order_data = {
        "order_date": order_date or datetime.now().strftime("%d/%m"),
        "customer_name": customer_name,
        "customer_phone": customer_phone.replace(" ", ""),
        "customer_address": customer_address,
        "source": "",
        "product_name": product_name,
        "weight": 0,
        "volume": 0,
        "tracking_cn": "",
        "tracking_vn": "",
        "account": "",
        "note": note,
        "total_price": total_price,
        "deposit": deposit,
        "remaining": 0,
        "status": "",
    }

    try:
        append_order_to_sheet(sheet_type, order_data)
        return HTMLResponse(f"""
        <div class="bg-green-50 border border-green-200 rounded-lg p-4">
            <div class="flex items-center gap-2 text-green-700 font-medium">
                <svg class="w-5 h-5" fill="currentColor" viewBox="0 0 20 20">
                    <path fill-rule="evenodd" d="M10 18a8 8 0 100-16 8 8 0 000 16zm3.707-9.293a1 1 0 00-1.414-1.414L9 10.586 7.707 9.293a1 1 0 00-1.414 1.414l2 2a1 1 0 001.414 0l4-4z" clip-rule="evenodd"/>
                </svg>
                Tạo đơn thành công!
            </div>
            <div class="text-sm text-green-600 mt-1">
                Đã ghi vào sheet {sheet_type} • {customer_name} • {total_price:,.0f} đ
            </div>
        </div>
        """)
    except Exception as e:
        # [H-08] Save to local buffer instead of just showing error
        save_to_buffer({"sheet_type": sheet_type, **order_data}, str(e))
        buffer_count = get_buffer_count()
        return HTMLResponse(f"""
        <div class="bg-amber-50 border border-amber-200 rounded-lg p-4">
            <div class="flex items-center gap-2 text-amber-700 font-medium">
                <svg class="w-5 h-5" fill="currentColor" viewBox="0 0 20 20">
                    <path fill-rule="evenodd" d="M8.257 3.099c.765-1.36 2.722-1.36 3.486 0l5.58 9.92c.75 1.334-.213 2.98-1.742 2.98H4.42c-1.53 0-2.493-1.646-1.743-2.98l5.58-9.92zM11 13a1 1 0 11-2 0 1 1 0 012 0zm-1-8a1 1 0 00-1 1v3a1 1 0 002 0V6a1 1 0 00-1-1z" clip-rule="evenodd"/>
                </svg>
                Không gửi được Google Sheets — Đã lưu tạm cục bộ
            </div>
            <div class="text-sm text-amber-600 mt-1">
                Đơn của <strong>{customer_name}</strong> đã lưu vào bộ đệm.
                {buffer_count} đơn đang chờ gửi. Sẽ tự động thử lại khi có kết nối.
            </div>
        </div>
        """)

@router.get("/calc-shipping")
async def calc_shipping_api(
    sheet: str = Query("DON", alias="sheet"),
    weight: float = Query(0, alias="weight"),
    volume: float = Query(0, alias="volume"),
):
    """Calculate shipping cost."""
    cost = calc_shipping(sheet, weight, volume)
    if sheet == "DON":
        formula = f"max({weight:.1f} × {DON_RATE_PER_KG:,}, {volume:.2f} × {DON_RATE_PER_M3:,})"
    else:
        formula = f"{weight:.1f} × {DON2_RATE_PER_KG:,}"

    return HTMLResponse(f"""
    <span class="text-sm font-medium text-blue-700">{cost:,.0f} đ</span>
    <span class="text-xs text-gray-500 ml-1">({formula})</span>
    """)

@router.get("/debt-summary")
async def debt_summary(q: str = Query("", alias="q")):
    """Debt summary for HTMX."""
    db = await get_db()
    try:
        like = f"%{q}%" if q else "%"
        debts = await db.execute_fetchall(
            """SELECT customer_name, customer_phone,
                COUNT(*) as order_count,
                COALESCE(SUM(total_price),0) as tp,
                COALESCE(SUM(deposit),0) as td,
                COALESCE(SUM(remaining),0) as tr
            FROM orders
            WHERE customer_name != '' AND remaining > 0
                AND (customer_name LIKE ? OR customer_phone LIKE ?)
            GROUP BY customer_phone
            ORDER BY tr DESC LIMIT 50""",
            (like, like)
        )

        rows = ""
        for d in debts:
            pct = (d[4] / d[3] * 100) if d[3] > 0 else 0
            bar_color = "red" if pct < 50 else "yellow" if pct < 80 else "green"
            rows += f"""
            <tr class="border-b hover:bg-gray-50">
                <td class="px-3 py-2 font-medium">{d[0]}</td>
                <td class="px-3 py-2 text-sm">{d[1]}</td>
                <td class="px-3 py-2 text-center">{d[2]}</td>
                <td class="px-3 py-2 text-right">{d[3]:,.0f} đ</td>
                <td class="px-3 py-2 text-right">{d[4]:,.0f} đ</td>
                <td class="px-3 py-2 text-right font-medium text-red-600">{d[5]:,.0f} đ</td>
                <td class="px-3 py-2">
                    <div class="w-full bg-gray-200 rounded-full h-2">
                        <div class="bg-{bar_color}-500 h-2 rounded-full" style="width:{pct:.0f}%"></div>
                    </div>
                </td>
            </tr>"""

        return HTMLResponse(rows)
    finally:
        await db.close()


# ==================== NEW API ENDPOINTS ====================

@router.get("/dashboard-data")
async def dashboard_data():
    """JSON data for dashboard charts."""
    db = await get_db()
    try:
        # Top 10 debt customers
        top_debt = await db.execute_fetchall(
            """SELECT customer_name, customer_phone,
                COALESCE(SUM(remaining),0) as debt,
                COUNT(*) as order_count
            FROM orders
            WHERE customer_name != '' AND remaining > 0
            GROUP BY customer_phone
            ORDER BY debt DESC
            LIMIT 10"""
        )

        # Shipping alerts: TQ tracking exists, no VN tracking, status not yet "Nhập kho TQ"
        # Shows orders that may be stuck in transit — user reviews order_date manually
        shipping_alerts = await db.execute_fetchall(
            """SELECT o.id, o.customer_name, o.customer_phone, o.tracking_cn,
                o.order_date, o.sheet_type, o.status,
                GROUP_CONCAT(oi.product_name, ' | ') as products
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            WHERE o.tracking_cn != '' AND o.tracking_cn IS NOT NULL
                AND (o.tracking_vn IS NULL OR o.tracking_vn = '')
                AND (o.status IS NULL OR o.status = '' OR o.status NOT LIKE '%Nhập kho TQ%')
                AND o.order_date != ''
            GROUP BY o.id
            ORDER BY o.order_date ASC
            LIMIT 50"""
        )

        # Overall stats
        stats = await db.execute_fetchall(
            """SELECT
                COUNT(*) as total_orders,
                COALESCE(SUM(total_price),0) as total_revenue,
                COALESCE(SUM(deposit),0) as total_collected,
                COALESCE(SUM(remaining),0) as total_debt,
                COUNT(CASE WHEN tracking_cn != '' AND (tracking_vn IS NULL OR tracking_vn = '') THEN 1 END) as pending_vn
            FROM orders"""
        )

        return JSONResponse({
            "top_debt": [{"name": r[0], "phone": r[1], "debt": r[2], "count": r[3]} for r in top_debt],
            "shipping_alerts": [
                {"id": r[0], "name": r[1], "phone": r[2], "tracking_cn": r[3],
                 "date": r[4], "sheet": r[5], "status": r[6], "products": r[7]}
                for r in shipping_alerts
            ],
            "stats": {
                "total_orders": stats[0][0] if stats else 0,
                "total_revenue": stats[0][1] if stats else 0,
                "total_collected": stats[0][2] if stats else 0,
                "total_debt": stats[0][3] if stats else 0,
                "pending_vn": stats[0][4] if stats else 0,
            }
        })
    finally:
        await db.close()


@router.get("/export-orders")
async def export_orders(
    sheet: str = Query("", alias="sheet"),
    status: str = Query("", alias="status"),
    date_from: str = Query("", alias="date_from"),
    date_to: str = Query("", alias="date_to"),
):
    """Export filtered orders to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = await get_db()
    try:
        where_clauses = []
        params = []
        if sheet:
            where_clauses.append("o.sheet_type = ?")
            params.append(sheet)
        if status:
            where_clauses.append("o.status = ?")
            params.append(status)

        where = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        rows = await db.execute_fetchall(
            f"""SELECT o.order_date, o.customer_name, o.customer_phone, o.customer_address,
                o.sheet_type, o.tracking_cn, o.tracking_vn,
                o.total_price, o.deposit, o.remaining, o.status, o.note,
                GROUP_CONCAT(oi.product_name, ' | ') as products,
                GROUP_CONCAT(oi.weight, ' | ') as weights,
                GROUP_CONCAT(oi.volume, ' | ') as volumes
            FROM orders o
            LEFT JOIN order_items oi ON oi.order_id = o.id
            {where}
            GROUP BY o.id
            ORDER BY o.row_start DESC""",
            params
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Đơn hàng"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["Ngày", "Khách hàng", "SĐT", "Địa chỉ", "Sheet", "Sản phẩm",
                    "VĐ TQ", "VĐ VN", "Tổng giá", "Cọc", "Còn lại", "Trạng thái", "Ghi chú"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, r in enumerate(rows, 2):
            values = [r[0], r[1], r[2], r[3], r[4], r[12], r[5], r[6],
                     r[7], r[8], r[9], r[10], r[11]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col in (10, 11, 12) and val:
                    cell.number_format = '#,##0'

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Summary row
        summary_row = len(rows) + 3
        ws.cell(row=summary_row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
        ws.cell(row=summary_row, column=9, value=sum(r[7] or 0 for r in rows)).font = Font(bold=True)
        ws.cell(row=summary_row, column=9).number_format = '#,##0'
        ws.cell(row=summary_row, column=10, value=sum(r[8] or 0 for r in rows)).font = Font(bold=True)
        ws.cell(row=summary_row, column=10).number_format = '#,##0'
        ws.cell(row=summary_row, column=11, value=sum(r[9] or 0 for r in rows)).font = Font(bold=True)
        ws.cell(row=summary_row, column=11).number_format = '#,##0'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"don-hang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        await db.close()


@router.post("/upload-credentials")
async def upload_credentials(file: UploadFile = File(...), user: str = Depends(verify_admin)):
    """Upload Google Service Account JSON credentials."""
    if not file.filename.endswith('.json'):
        return HTMLResponse("""
        <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
            ❌ File phải có đuôi .json
        </div>
        """, status_code=400)

    try:
        content = await file.read()
        data = json.loads(content)

        if "type" not in data or data.get("type") != "service_account":
            return HTMLResponse("""
            <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
                ❌ File không phải Service Account credentials hợp lệ
            </div>
            """, status_code=400)

        dest = CREDENTIALS_DIR / file.filename
        with open(dest, "wb") as f:
            f.write(content)

        env_path = BASE_DIR / ".env"
        _update_env_var(env_path, "GOOGLE_CREDS_FILE", str(dest))

        return HTMLResponse(f"""
        <div class="bg-green-50 border border-green-200 rounded-lg p-3 text-green-700 text-sm">
            ✅ Đã upload <strong>{file.filename}</strong> vào thư mục credentials/
            <br><span class="text-green-600">Đã cập nhật GOOGLE_CREDS_FILE trong .env. Restart server để áp dụng.</span>
        </div>
        """)
    except json.JSONDecodeError:
        return HTMLResponse("""
        <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
            ❌ File không phải JSON hợp lệ
        </div>
        """, status_code=400)
    except Exception as e:
        return HTMLResponse(f"""
        <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
            ❌ Lỗi: {str(e)}
        </div>
        """, status_code=500)


@router.post("/update-config")
async def update_config(
    GOOGLE_CREDS_FILE: str = Form(""),
    SPREADSHEET_ID: str = Form(""),
    GHN_TOKEN: str = Form(""),
    VTP_TOKEN: str = Form(""),
    user: str = Depends(verify_admin),
):
    """Update .env configuration file."""
    try:
        env_path = BASE_DIR / ".env"

        if GOOGLE_CREDS_FILE:
            _update_env_var(env_path, "GOOGLE_CREDS_FILE", GOOGLE_CREDS_FILE)
        if SPREADSHEET_ID:
            _update_env_var(env_path, "SPREADSHEET_ID", SPREADSHEET_ID)
        _update_env_var(env_path, "GHN_TOKEN", GHN_TOKEN)
        _update_env_var(env_path, "VTP_TOKEN", VTP_TOKEN)

        return HTMLResponse("""
        <div class="bg-green-50 border border-green-200 rounded-lg p-3 text-green-700 text-sm">
            ✅ Đã lưu cấu hình vào .env. <strong>Restart server để áp dụng thay đổi.</strong>
        </div>
        """)
    except Exception as e:
        return HTMLResponse(f"""
        <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
            ❌ Lỗi: {str(e)}
        </div>
        """, status_code=500)


def _update_env_var(env_path: str, key: str, value: str):
    """Update or add a key=value in .env file."""
    lines = []
    found = False

    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            lines = f.readlines()

    new_lines = []
    for line in lines:
        if line.strip().startswith(f"{key}="):
            new_lines.append(f"{key}={value}\n")
            found = True
        else:
            new_lines.append(line)

    if not found:
        new_lines.append(f"{key}={value}\n")

    with open(env_path, "w") as f:
        f.writelines(new_lines)


@router.post("/flush-buffer")
async def flush_buffer_endpoint(user: str = Depends(verify_user)):
    """Retry sending buffered orders to Google Sheets."""
    try:
        sent, failed = await flush_buffer()
        if sent > 0 and failed == 0:
            return HTMLResponse(f"""
            <div class="bg-green-50 border border-green-200 rounded-lg p-3 text-green-700 text-sm">
                ✅ Đã gửi thành công {sent} đơn từ bộ đệm
            </div>
            """)
        elif sent > 0:
            return HTMLResponse(f"""
            <div class="bg-yellow-50 border border-yellow-200 rounded-lg p-3 text-yellow-700 text-sm">
                ⚠️ Gửi {sent} đơn thành công, {failed} đơn vẫn lỗi
            </div>
            """)
        else:
            return HTMLResponse("""
            <div class="bg-gray-50 border border-gray-200 rounded-lg p-3 text-gray-600 text-sm">
                📭 Bộ đệm trống — không có đơn nào cần gửi
            </div>
            """)
    except Exception as e:
        return HTMLResponse(f"""
        <div class="bg-red-50 border border-red-200 rounded-lg p-3 text-red-700 text-sm">
            ❌ Lỗi: {str(e)}
        </div>
        """, status_code=500)
